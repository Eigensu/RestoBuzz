"""
ReserveGo upload portal.

Auth: simple username/password against env vars RESERVEGO_USER / RESERVEGO_PASSWORD.
Returns a short-lived JWT (role=reservego) used to authenticate the upload endpoint.

Performance: openpyxl parsing (CPU-bound) runs in a thread pool; DB writes use
bulk_write so the entire sheet is one round-trip instead of N awaits.
"""

import io
import asyncio
import re
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timezone, timedelta
from typing import Annotated

import openpyxl
from fastapi import APIRouter, Body, Depends, UploadFile, File, Query
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from motor.motor_asyncio import AsyncIOMotorDatabase
from pydantic import BaseModel
from pymongo import UpdateOne

from app.config import settings
from app.core.errors import InvalidCredentialsError, InvalidFileFormatError, AppError
from app.core.security import _create_token, decode_token
from app.database import get_db
from app.dependencies import require_role, validate_restaurant_access

router = APIRouter(prefix="/reservego", tags=["reservego"])
_bearer = HTTPBearer()
_executor = ThreadPoolExecutor(max_workers=2)

# ── Shared Constants (SonarCloud Hardening) ───────────────────────────────────
_MONGO_GROUP = "$group"
_MONGO_LIMIT = "$limit"
_MONGO_MATCH = "$match"
_MONGO_SORT = "$sort"
_MONGO_SUM = "$sum"
_MONGO_AVG = "$avg"
_MONGO_YEAR = "$year"
_MONGO_MONTH = "$month"
_MONGO_BUCKET = "$bucket"
_MONGO_FIRST = "$first"
_MONGO_LOOKUP = "$lookup"
_MONGO_REGEX = "$regex"
_MONGO_OPTIONS = "$options"
_MONGO_GT = "$gt"
_MONGO_NE = "$ne"

_COL_PHONE_NUMBER = "phone number"
_COL_EMAIL_ID = "email id"
_COL_TOTAL_VISITS = "total visits"
_COL_GUEST_NAME = "guest name"
_COL_BILL_AMOUNT = "bill amount"
_COL_BILL_NUMBER = "bill number"
_COL_BOOKING_TIME = "booking time"
_COL_GUEST_NUMBER = "guest number"
_COL_GUEST_EMAIL = "guest email"
_COL_LAST_VISITED = "last visited date"

_FLD_BILL_AMOUNT = "$bill_amount"


class _PortalNotConfiguredError(AppError):
    status_code = 503
    error_type = "not_configured"


GUEST_PROFILE_HEADERS = [
    _COL_GUEST_NAME,
    _COL_PHONE_NUMBER,
    _COL_EMAIL_ID,
    _COL_TOTAL_VISITS,
    "source",
    "mode",
    _COL_LAST_VISITED,
    "birthday",
    "anniversary",
]

BILL_HEADERS = [
    "sno",
    "outlet name",
    _COL_BOOKING_TIME,
    "seated time",
    "reserved time",
    "booking type",
    _COL_GUEST_NAME,
    _COL_GUEST_NUMBER,
    _COL_GUEST_EMAIL,
    "pax",
    "reserved by",
    "section(s)",
    "table(s)",
    "vist count",
    "booking status",
    "deletion type",
    "deleted reason",
    "source of booking",
    "preferences",
    "tags",
    "guest comments",
    "outlet comments",
    _COL_BILL_AMOUNT,
    _COL_BILL_NUMBER,
    "booking amount",
    "booking amount tranx id",
    "booking amount payment status",
    "booking amount payment date",
]

GUEST_PROFILE_SHEETS = {"no of visits", "no visit data", "not visited in 3 months"}
BILL_SHEET = _COL_BILL_AMOUNT


# ── Auth ──────────────────────────────────────────────────────────────────────


def _mint_token() -> str:
    return _create_token(
        {"sub": "reservego", "role": "reservego", "type": "access"},
        timedelta(hours=6),
    )


def _require_token(
    credentials: HTTPAuthorizationCredentials = Depends(_bearer),
) -> None:
    try:
        payload = decode_token(credentials.credentials)
    except ValueError as exc:
        raise InvalidCredentialsError("Invalid or expired token") from exc
    if payload.get("role") != "reservego":
        raise InvalidCredentialsError("Not a reservego token")


# ── Schemas ───────────────────────────────────────────────────────────────────


class LoginRequest(BaseModel):
    username: str
    password: str


class LoginResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"


# ── Helpers ───────────────────────────────────────────────────────────────────


# Aliases map canonical header names → alternative names that may appear in the Excel
_GUEST_HEADER_ALIASES: dict[str, list[str]] = {
    "phone number": ["guest number"],
    "email id": ["guest email"],
    "total visits": ["visits"],
    "source": ["registration source"],
    "mode": ["registration mode"],
    "last visited date": ["last visit"],
}


def _make_idx(raw_headers: list, headers: list) -> dict:
    """Build a column-index map, falling back to known aliases when the canonical
    header is absent from the sheet."""
    idx: dict[str, int | None] = {}
    for h in headers:
        if h in raw_headers:
            idx[h] = raw_headers.index(h)
        else:
            # Try each alias in order
            resolved = None
            for alias in _GUEST_HEADER_ALIASES.get(h, []):
                if alias in raw_headers:
                    resolved = raw_headers.index(alias)
                    break
            idx[h] = resolved
    return idx


def _cell(row: tuple, idx: dict, h: str):
    i = idx.get(h)
    return row[i] if i is not None and i < len(row) else None


def _parse_date(val) -> datetime | None:
    return val.replace(tzinfo=timezone.utc) if isinstance(val, datetime) else None


def _str_val(row, idx, h):
    v = _cell(row, idx, h)
    s = str(v).strip() if v is not None else None
    return None if not s or s.lower() == "none" else s


def _num_val(row, idx, h):
    v = _cell(row, idx, h)
    try:
        return float(v) if v is not None else None
    except (ValueError, TypeError):
        return None


def _sheet_headers(ws) -> list:
    return [
        str(c).strip().lower() if c is not None else ""
        for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
    ]


# ── Sync parsers (run in thread) ──────────────────────────────────────────────


def _parse_sheet_rows(ws, parse_fn, idx) -> tuple[list, int]:
    """Generic row iterator — returns (docs, skipped)."""
    docs, skipped = [], 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        doc = parse_fn(row, idx)
        if doc is None:
            skipped += 1
        else:
            docs.append(doc)
    return docs, skipped


def _parse_workbook(
    contents: bytes, filename: str, now: datetime, restaurant_id: str
) -> dict:
    """Parse all sheets — runs synchronously, call via run_in_executor."""
    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True, read_only=True)
    result = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        key = sheet_name.strip().lower()
        raw_headers = _sheet_headers(ws)

        if key in GUEST_PROFILE_SHEETS:
            idx = _make_idx(raw_headers, GUEST_PROFILE_HEADERS)
            docs, skipped = _parse_sheet_rows(
                ws,
                lambda row, i, sn=sheet_name: _parse_guest_row(
                    row, i, sn, filename, now, restaurant_id
                ),
                idx,
            )
            result[sheet_name] = ("guest", docs, skipped)

        elif key == BILL_SHEET:
            idx = _make_idx(raw_headers, BILL_HEADERS)
            docs, skipped = _parse_sheet_rows(
                ws,
                lambda row, i: _parse_bill_row(row, i, filename, now, restaurant_id),
                idx,
            )
            result[sheet_name] = ("bill", docs, skipped)

        else:
            result[sheet_name] = ("unknown", [], 0)

    wb.close()
    return result


def _parse_guest_row(row, idx, sheet_name, filename, now, restaurant_id) -> dict | None:
    guest_name = str(_cell(row, idx, _COL_GUEST_NAME) or "").strip()
    if not guest_name or guest_name.lower() == "none":
        return None

    raw_phone = _cell(row, idx, _COL_PHONE_NUMBER)
    phone = ""
    if raw_phone is not None:
        phone = (
            str(int(raw_phone))
            if isinstance(raw_phone, float)
            else str(raw_phone).strip()
        )
        phone = phone.replace(" ", "").replace("+", "").replace("-", "")

    email_val = _cell(row, idx, _COL_EMAIL_ID)
    email = str(email_val).strip() if email_val else None
    if email and email.lower() == "none":
        email = None

    total_visits_val = _cell(row, idx, _COL_TOTAL_VISITS)
    total_visits = int(total_visits_val) if total_visits_val is not None else 0

    return {
        "guest_name": guest_name,
        "phone": phone,
        "email": email,
        "total_visits": total_visits,
        "source": str(_cell(row, idx, "source") or "").strip() or None,
        "mode": str(_cell(row, idx, "mode") or "").strip() or None,
        "last_visited_date": _parse_date(_cell(row, idx, _COL_LAST_VISITED)),
        "birthday": _parse_date(_cell(row, idx, "birthday")),
        "anniversary": _parse_date(_cell(row, idx, "anniversary")),
        "sheet": sheet_name,
        "restaurant_id": restaurant_id,
        "uploaded_at": now,
        "filename": filename,
    }


def _parse_bill_row(row, idx, filename, now, restaurant_id) -> dict | None:
    guest_name = str(_cell(row, idx, _COL_GUEST_NAME) or "").strip()
    if not guest_name or guest_name.lower() == "none":
        return None

    return {
        "sno": _num_val(row, idx, "sno"),
        "outlet_name": _str_val(row, idx, "outlet name"),
        "booking_time": _parse_date(_cell(row, idx, _COL_BOOKING_TIME)),
        "seated_time": _parse_date(_cell(row, idx, "seated time")),
        "reserved_time": _parse_date(_cell(row, idx, "reserved time")),
        "booking_type": _str_val(row, idx, "booking type"),
        "guest_name": guest_name,
        "guest_number": _str_val(row, idx, _COL_GUEST_NUMBER),
        "guest_email": _str_val(row, idx, _COL_GUEST_EMAIL),
        "pax": _num_val(row, idx, "pax"),
        "reserved_by": _str_val(row, idx, "reserved by"),
        "sections": _str_val(row, idx, "section(s)"),
        "tables": _str_val(row, idx, "table(s)"),
        "visit_count": _num_val(row, idx, "vist count"),
        "booking_status": _str_val(row, idx, "booking status"),
        "deletion_type": _str_val(row, idx, "deletion type"),
        "deleted_reason": _str_val(row, idx, "deleted reason"),
        "source_of_booking": _str_val(row, idx, "source of booking"),
        "preferences": _str_val(row, idx, "preferences"),
        "tags": _str_val(row, idx, "tags"),
        "guest_comments": _str_val(row, idx, "guest comments"),
        "outlet_comments": _str_val(row, idx, "outlet comments"),
        "bill_amount": _num_val(row, idx, _COL_BILL_AMOUNT),
        "bill_number": _str_val(row, idx, _COL_BILL_NUMBER),
        "booking_amount": _num_val(row, idx, "booking amount"),
        "booking_amount_tranx_id": _str_val(row, idx, "booking amount tranx id"),
        "booking_amount_payment_status": _str_val(
            row, idx, "booking amount payment status"
        ),
        "booking_amount_payment_date": _parse_date(
            _cell(row, idx, "booking amount payment date")
        ),
        "restaurant_id": restaurant_id,
        "uploaded_at": now,
        "filename": filename,
    }


# ── Bulk DB writers ───────────────────────────────────────────────────────────


async def _bulk_upsert_guests(docs: list, db: AsyncIOMotorDatabase) -> None:
    if not docs:
        return
    ops = []
    for doc in docs:
        if doc["phone"]:
            ops.append(
                UpdateOne(
                    {"phone": doc["phone"], "restaurant_id": doc["restaurant_id"]},
                    {"$set": doc},
                    upsert=True,
                )
            )
        else:
            ops.append(
                UpdateOne(
                    {
                        "guest_name": doc["guest_name"],
                        "email": doc["email"],
                        "sheet": doc["sheet"],
                        "restaurant_id": doc["restaurant_id"],
                    },
                    {"$set": doc},
                    upsert=True,
                )
            )
    await db.reservego_uploads.bulk_write(ops, ordered=False)


async def _bulk_upsert_bills(docs: list, db: AsyncIOMotorDatabase) -> None:
    if not docs:
        return
    ops = []
    for doc in docs:
        if doc.get("bill_number"):
            ops.append(
                UpdateOne(
                    {
                        "bill_number": doc["bill_number"],
                        "restaurant_id": doc["restaurant_id"],
                    },
                    {"$set": doc},
                    upsert=True,
                )
            )
        else:
            ops.append(
                UpdateOne(
                    {
                        "guest_name": doc["guest_name"],
                        "booking_time": doc["booking_time"],
                        "restaurant_id": doc["restaurant_id"],
                    },
                    {"$set": doc},
                    upsert=True,
                )
            )
    await db.reservego_bill_data.bulk_write(ops, ordered=False)


# ── Routes ────────────────────────────────────────────────────────────────────


@router.post("/login", response_model=LoginResponse)
async def reservego_login(body: Annotated[LoginRequest, Body()]):
    if not settings.reservego_user or not settings.reservego_password:
        raise _PortalNotConfiguredError("ReserveGo portal not configured")
    if (
        body.username != settings.reservego_user
        or body.password != settings.reservego_password
    ):
        raise InvalidCredentialsError("Invalid username or password")
    return LoginResponse(access_token=_mint_token())


class Restaurant(BaseModel):
    id: str
    name: str


@router.get("/restaurants", response_model=list[Restaurant])
async def list_restaurants(
    _auth: Annotated[None, Depends(_require_token)],
    db: Annotated[AsyncIOMotorDatabase, Depends(get_db)],
):
    """Return all restaurants so the upload portal can show a selector."""
    cursor = db.restaurants.find({}, {"id": 1, "name": 1})
    return [
        Restaurant(id=str(doc.get("id") or doc["_id"]), name=doc["name"])
        async for doc in cursor
    ]


def _serialize_doc(doc: dict) -> dict:
    """Convert ObjectId and datetime fields to JSON-safe types."""
    result = {}
    for k, v in doc.items():
        if isinstance(v, datetime):
            result[k] = v.isoformat()
        else:
            result[k] = v
    return result


@router.get("/analytics")
async def get_analytics(
    restaurant_id: Annotated[str, Query()],
    current_user: Annotated[dict, Depends(require_role("viewer"))] = None,
    db: Annotated[AsyncIOMotorDatabase, Depends(get_db)] = None,
):
    # Security: Validate restaurant access
    await validate_restaurant_access(current_user, restaurant_id, db)

    # ── Guest stats ───────────────────────────────────────────────────────────
    total_guests = await db.reservego_uploads.count_documents(
        {"restaurant_id": restaurant_id}
    )
    with_phone = await db.reservego_uploads.count_documents(
        {"restaurant_id": restaurant_id, "phone": {"$nin": ["", None]}}
    )
    with_email = await db.reservego_uploads.count_documents(
        {"restaurant_id": restaurant_id, "email": {"$nin": ["", None]}}
    )

    # ── Bill stats ────────────────────────────────────────────────────────────
    total_bills = await db.reservego_bill_data.count_documents(
        {"restaurant_id": restaurant_id}
    )

    rev_pipeline = [
        {_MONGO_MATCH: {"restaurant_id": restaurant_id, "bill_amount": {_MONGO_GT: 0}}},
        {
            _MONGO_GROUP: {
                "_id": None,
                "total": {_MONGO_SUM: _FLD_BILL_AMOUNT},
                "avg": {_MONGO_AVG: _FLD_BILL_AMOUNT},
                "count": {_MONGO_SUM: 1},
            }
        },
    ]
    rev_result = await db.reservego_bill_data.aggregate(rev_pipeline).to_list(1)
    total_revenue = rev_result[0]["total"] if rev_result else 0
    avg_bill = rev_result[0]["avg"] if rev_result else 0
    bills_with_amount = rev_result[0]["count"] if rev_result else 0

    # ── Monthly revenue trend ─────────────────────────────────────────────────
    monthly_pipeline = [
        {
            _MONGO_MATCH: {
                "restaurant_id": restaurant_id,
                "bill_amount": {_MONGO_GT: 0},
                "booking_time": {_MONGO_NE: None},
            }
        },
        {
            _MONGO_GROUP: {
                "_id": {
                    "year": {_MONGO_YEAR: "$booking_time"},
                    "month": {_MONGO_MONTH: "$booking_time"},
                },
                "revenue": {_MONGO_SUM: _FLD_BILL_AMOUNT},
                "bookings": {_MONGO_SUM: 1},
                "avg_pax": {_MONGO_AVG: "$pax"},
            }
        },
        {_MONGO_SORT: {"_id.year": -1, "_id.month": -1}},
    ]
    monthly_raw = await db.reservego_bill_data.aggregate(monthly_pipeline).to_list(24)
    # Reverse so the chart shows oldest to newest (up to last 24 months)
    monthly_raw.reverse()

    month_names = [
        "Jan",
        "Feb",
        "Mar",
        "Apr",
        "May",
        "Jun",
        "Jul",
        "Aug",
        "Sep",
        "Oct",
        "Nov",
        "Dec",
    ]
    monthly_trend = [
        {
            "month": f"{month_names[r['_id']['month'] - 1]} {r['_id']['year']}",
            "revenue": round(r["revenue"]),
            "bookings": r["bookings"],
            "avg_pax": round(r["avg_pax"] or 0, 1),
        }
        for r in monthly_raw
    ]

    # ── Booking status breakdown ──────────────────────────────────────────────
    status_pipeline = [
        {
            _MONGO_MATCH: {
                "restaurant_id": restaurant_id,
                "booking_status": {"$nin": [None, ""]},
            }
        },
        {_MONGO_GROUP: {"_id": "$booking_status", "count": {_MONGO_SUM: 1}}},
        {_MONGO_SORT: {"count": -1}},
    ]
    status_raw = await db.reservego_bill_data.aggregate(status_pipeline).to_list(20)
    booking_statuses = [{"status": r["_id"], "count": r["count"]} for r in status_raw]

    # ── Booking type (walkin vs reservation) ──────────────────────────────────
    type_pipeline = [
        {
            _MONGO_MATCH: {
                "restaurant_id": restaurant_id,
                "booking_type": {"$nin": [None, ""]},
            }
        },
        {_MONGO_GROUP: {"_id": "$booking_type", "count": {_MONGO_SUM: 1}}},
    ]
    type_raw = await db.reservego_bill_data.aggregate(type_pipeline).to_list(10)
    booking_types = [{"type": r["_id"], "count": r["count"]} for r in type_raw]

    # ── Source of booking ─────────────────────────────────────────────────────
    source_pipeline = [
        {
            _MONGO_MATCH: {
                "restaurant_id": restaurant_id,
                "source_of_booking": {"$nin": [None, ""]},
            }
        },
        {
            _MONGO_GROUP: {
                "_id": "$source_of_booking",
                "count": {_MONGO_SUM: 1},
                "revenue": {_MONGO_SUM: _FLD_BILL_AMOUNT},
            }
        },
        {_MONGO_SORT: {"count": -1}},
        {_MONGO_LIMIT: 8},
    ]
    source_raw = await db.reservego_bill_data.aggregate(source_pipeline).to_list(8)
    booking_sources = [
        {"source": r["_id"], "count": r["count"], "revenue": round(r["revenue"] or 0)}
        for r in source_raw
    ]

    # ── Top sections by revenue ───────────────────────────────────────────────
    section_pipeline = [
        {
            _MONGO_MATCH: {
                "restaurant_id": restaurant_id,
                "sections": {"$nin": [None, ""]},
                "bill_amount": {_MONGO_GT: 0},
            }
        },
        {
            _MONGO_GROUP: {
                "_id": "$sections",
                "revenue": {_MONGO_SUM: _FLD_BILL_AMOUNT},
                "count": {_MONGO_SUM: 1},
            }
        },
        {_MONGO_SORT: {"revenue": -1}},
        {_MONGO_LIMIT: 6},
    ]
    section_raw = await db.reservego_bill_data.aggregate(section_pipeline).to_list(6)
    top_sections = [
        {"section": r["_id"], "count": r["count"], "revenue": round(r["revenue"])}
        for r in section_raw
    ]

    # ── Guest visit frequency distribution ───────────────────────────────────
    visit_pipeline = [
        {_MONGO_MATCH: {"restaurant_id": restaurant_id}},
        {
            _MONGO_BUCKET: {
                "groupBy": "$total_visits",
                "boundaries": [0, 1, 2, 3, 5, 10, 20],
                "default": "20+",
                "output": {"count": {_MONGO_SUM: 1}},
            }
        },
    ]
    visit_raw = await db.reservego_uploads.aggregate(visit_pipeline).to_list(10)
    visit_labels = {
        0: "0 visits",
        1: "1 visit",
        2: "2 visits",
        3: "3-4 visits",
        5: "5-9 visits",
        10: "10-19 visits",
        "20+": "20+ visits",
    }
    visit_dist = [
        {"label": visit_labels.get(r["_id"], str(r["_id"])), "count": r["count"]}
        for r in visit_raw
    ]

    # ── Guest source breakdown (from uploads) ─────────────────────────────────
    guest_source_pipeline = [
        {_MONGO_MATCH: {"restaurant_id": restaurant_id, "source": {"$nin": [None, ""]}}},
        {_MONGO_GROUP: {"_id": "$source", "count": {_MONGO_SUM: 1}}},
        {_MONGO_SORT: {"count": -1}},
        {_MONGO_LIMIT: 6},
    ]
    guest_source_raw = await db.reservego_uploads.aggregate(
        guest_source_pipeline
    ).to_list(6)
    guest_sources = [
        {"source": r["_id"], "count": r["count"]} for r in guest_source_raw
    ]

    return {
        "summary": {
            "total_guests": total_guests,
            "with_phone": with_phone,
            "with_email": with_email,
            "total_bills": total_bills,
            "total_revenue": round(total_revenue),
            "avg_bill": round(avg_bill),
            "bills_with_amount": bills_with_amount,
        },
        "monthly_trend": monthly_trend,
        "booking_statuses": booking_statuses,
        "booking_types": booking_types,
        "booking_sources": booking_sources,
        "top_sections": top_sections,
        "visit_distribution": visit_dist,
        "guest_sources": guest_sources,
    }


@router.get("/guests")
async def list_guests(
    restaurant_id: Annotated[str, Query()],
    sheet: Annotated[str | None, Query()] = None,
    search: Annotated[str | None, Query()] = None,
    page: Annotated[int, Query(ge=1)] = 1,
    page_size: Annotated[int, Query(ge=1, le=200)] = 50,
    current_user: Annotated[dict, Depends(require_role("viewer"))] = None,
    db: Annotated[AsyncIOMotorDatabase, Depends(get_db)] = None,
):
    # Security: Validate restaurant access
    await validate_restaurant_access(current_user, restaurant_id, db)

    query: dict = {"restaurant_id": restaurant_id}
    if sheet:
        query["sheet"] = {_MONGO_REGEX: re.escape(sheet), _MONGO_OPTIONS: "i"}
    if search:
        safe_search = re.escape(search)
        query["$or"] = [
            {"guest_name": {_MONGO_REGEX: safe_search, _MONGO_OPTIONS: "i"}},
            {"phone": {_MONGO_REGEX: safe_search, _MONGO_OPTIONS: "i"}},
            {"email": {_MONGO_REGEX: safe_search, _MONGO_OPTIONS: "i"}},
        ]
    skip = (page - 1) * page_size
    total = await db.reservego_uploads.count_documents(query)
    cursor = (
        db.reservego_uploads.find(query)
        .sort([("total_visits", -1), ("uploaded_at", -1)])
        .skip(skip)
        .limit(page_size)
    )
    items = []
    async for doc in cursor:
        doc["id"] = str(doc["_id"])
        doc.pop("_id", None)
        items.append(_serialize_doc(doc))
    return {"items": items, "total": total, "page": page, "page_size": page_size}


@router.get("/guests/sheets")
async def list_guest_sheets(
    restaurant_id: Annotated[str, Query()],
    current_user: Annotated[dict, Depends(require_role("viewer"))] = None,
    db: Annotated[AsyncIOMotorDatabase, Depends(get_db)] = None,
):
    # Security: Validate restaurant access
    await validate_restaurant_access(current_user, restaurant_id, db)

    sheets = await db.reservego_uploads.distinct(
        "sheet", {"restaurant_id": restaurant_id}
    )
    return {"sheets": sheets}


@router.get("/bills")
async def list_bills(
    restaurant_id: Annotated[str, Query()],
    search: Annotated[str | None, Query()] = None,
    page: Annotated[int, Query(ge=1)] = 1,
    page_size: Annotated[int, Query(ge=1, le=200)] = 50,
    current_user: Annotated[dict, Depends(require_role("viewer"))] = None,
    db: Annotated[AsyncIOMotorDatabase, Depends(get_db)] = None,
):
    # Security: Validate restaurant access
    await validate_restaurant_access(current_user, restaurant_id, db)

    query: dict = {"restaurant_id": restaurant_id}
    if search:
        safe_search = re.escape(search)
        query["$or"] = [
            {"guest_name": {_MONGO_REGEX: safe_search, _MONGO_OPTIONS: "i"}},
            {"guest_number": {_MONGO_REGEX: safe_search, _MONGO_OPTIONS: "i"}},
            {"bill_number": {_MONGO_REGEX: safe_search, _MONGO_OPTIONS: "i"}},
        ]
    skip = (page - 1) * page_size
    total = await db.reservego_bill_data.count_documents(query)
    cursor = (
        db.reservego_bill_data.find(query)
        .sort("booking_time", -1)
        .skip(skip)
        .limit(page_size)
    )
    items = []
    async for doc in cursor:
        doc["id"] = str(doc["_id"])
        doc.pop("_id", None)
        items.append(_serialize_doc(doc))
    return {"items": items, "total": total, "page": page, "page_size": page_size}


def _fmt_dt(val) -> str:
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d %H:%M")
    return str(val)


def _build_guests_wb(docs: list) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Guests"
    headers = [
        "Guest Name",
        "Phone",
        "Email",
        "Total Visits",
        "Source",
        "Mode",
        "Last Visited Date",
        "Birthday",
        "Anniversary",
        "Sheet",
        "Uploaded At",
        "Filename",
    ]
    ws.append(headers)
    for doc in docs:
        ws.append(
            [
                doc.get("guest_name", ""),
                doc.get("phone", ""),
                doc.get("email", ""),
                doc.get("total_visits", 0),
                doc.get("source", ""),
                doc.get("mode", ""),
                _fmt_dt(doc.get("last_visited_date")),
                _fmt_dt(doc.get("birthday")),
                _fmt_dt(doc.get("anniversary")),
                doc.get("sheet", ""),
                _fmt_dt(doc.get("uploaded_at")),
                doc.get("filename", ""),
            ]
        )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_bills_wb(docs: list) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bills"
    headers = [
        "Guest Name",
        "Guest Number",
        "Guest Email",
        "Outlet",
        "Booking Time",
        "Seated Time",
        "Reserved Time",
        "Booking Type",
        "Pax",
        "Reserved By",
        "Section(s)",
        "Table(s)",
        "Visit Count",
        "Booking Status",
        "Bill Amount",
        "Bill Number",
        "Booking Amount",
        "Source of Booking",
        "Tags",
        "Guest Comments",
        "Outlet Comments",
        "Uploaded At",
        "Filename",
    ]
    ws.append(headers)
    for doc in docs:
        ws.append(
            [
                doc.get("guest_name", ""),
                doc.get("guest_number", ""),
                doc.get("guest_email", ""),
                doc.get("outlet_name", ""),
                _fmt_dt(doc.get("booking_time")),
                _fmt_dt(doc.get("seated_time")),
                _fmt_dt(doc.get("reserved_time")),
                doc.get("booking_type", ""),
                doc.get("pax"),
                doc.get("reserved_by", ""),
                doc.get("sections", ""),
                doc.get("tables", ""),
                doc.get("visit_count"),
                doc.get("booking_status", ""),
                doc.get("bill_amount"),
                doc.get("bill_number", ""),
                doc.get("booking_amount"),
                doc.get("source_of_booking", ""),
                doc.get("tags", ""),
                doc.get("guest_comments", ""),
                doc.get("outlet_comments", ""),
                _fmt_dt(doc.get("uploaded_at")),
                doc.get("filename", ""),
            ]
        )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/guests/export")
async def export_guests(
    restaurant_id: Annotated[str, Query()],
    sheet: Annotated[str | None, Query()] = None,
    search: Annotated[str | None, Query()] = None,
    current_user: Annotated[dict, Depends(require_role("viewer"))] = None,
    db: Annotated[AsyncIOMotorDatabase, Depends(get_db)] = None,
):
    """Export all matching guest profiles as an xlsx file."""
    from fastapi.responses import Response

    # Security: Validate restaurant access
    await validate_restaurant_access(current_user, restaurant_id, db)

    query: dict = {"restaurant_id": restaurant_id}
    if sheet:
        query["sheet"] = {_MONGO_REGEX: re.escape(sheet), _MONGO_OPTIONS: "i"}
    if search:
        safe_search = re.escape(search)
        query["$or"] = [
            {"guest_name": {_MONGO_REGEX: safe_search, _MONGO_OPTIONS: "i"}},
            {"phone": {_MONGO_REGEX: safe_search, _MONGO_OPTIONS: "i"}},
            {"email": {_MONGO_REGEX: safe_search, _MONGO_OPTIONS: "i"}},
        ]

    # Optimization: Hard limit for export to avoid OOM
    MAX_ROWS = 50000
    cursor = (
        db.reservego_uploads.find(query)
        .sort([("total_visits", -1), ("uploaded_at", -1)])
        .limit(MAX_ROWS)
    )
    docs = [doc async for doc in cursor]

    loop = asyncio.get_running_loop()
    xlsx_bytes = await loop.run_in_executor(_executor, _build_guests_wb, docs)

    slug = (sheet or "guests").lower().replace(" ", "_")
    filename = f"reservego_{slug}.xlsx"
    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/bills/export")
async def export_bills(
    restaurant_id: Annotated[str, Query()],
    search: Annotated[str | None, Query()] = None,
    current_user: Annotated[dict, Depends(require_role("viewer"))] = None,
    db: Annotated[AsyncIOMotorDatabase, Depends(get_db)] = None,
):
    """Export all matching bill records as an xlsx file."""
    from fastapi.responses import Response

    # Security: Validate restaurant access
    await validate_restaurant_access(current_user, restaurant_id, db)

    query: dict = {"restaurant_id": restaurant_id}
    if search:
        safe_search = re.escape(search)
        query["$or"] = [
            {"guest_name": {_MONGO_REGEX: safe_search, _MONGO_OPTIONS: "i"}},
            {"guest_number": {_MONGO_REGEX: safe_search, _MONGO_OPTIONS: "i"}},
            {"bill_number": {_MONGO_REGEX: safe_search, _MONGO_OPTIONS: "i"}},
        ]

    # Optimization: Hard limit for export to avoid OOM
    MAX_ROWS = 50000
    docs = [
        doc
        async for doc in db.reservego_bill_data.find(query)
        .sort("booking_time", -1)
        .limit(MAX_ROWS)
    ]

    loop = asyncio.get_running_loop()
    xlsx_bytes = await loop.run_in_executor(_executor, _build_bills_wb, docs)

    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="reservego_bills.xlsx"'},
    )


@router.post("/upload")
async def reservego_upload(
    file: Annotated[UploadFile, File()],
    _auth: Annotated[None, Depends(_require_token)],
    db: Annotated[AsyncIOMotorDatabase, Depends(get_db)],
    restaurant_id: Annotated[str, Query()],
    current_user: Annotated[dict, Depends(require_role("viewer"))] = None,
    data_from: Annotated[str | None, Query()] = None,
    data_until: Annotated[str | None, Query()] = None,
):
    # Security: Validate restaurant access
    await validate_restaurant_access(current_user, restaurant_id, db)

    filename = file.filename or ""
    if not filename.lower().endswith(".xlsx"):
        raise InvalidFileFormatError("Only .xlsx Excel files are supported")

    contents = await file.read()
    if not contents:
        raise InvalidFileFormatError("Uploaded file is empty")

    now = datetime.now(timezone.utc)

    loop = asyncio.get_running_loop()
    try:
        parsed = await loop.run_in_executor(
            _executor, _parse_workbook, contents, filename, now, restaurant_id
        )
    except Exception as exc:
        raise InvalidFileFormatError("Unable to read Excel file") from exc

    sheets_summary = {}
    for sheet_name, (kind, docs, skipped) in parsed.items():
        # Stamp data_from / data_until on every doc
        if data_from or data_until:
            for doc in docs:
                doc["data_from"] = data_from
                doc["data_until"] = data_until
        if kind == "guest":
            await _bulk_upsert_guests(docs, db)
            sheets_summary[sheet_name] = {"inserted": len(docs), "skipped": skipped}
        elif kind == "bill":
            await _bulk_upsert_bills(docs, db)
            sheets_summary[sheet_name] = {"inserted": len(docs), "skipped": skipped}
        else:
            sheets_summary[sheet_name] = {
                "inserted": 0,
                "skipped": 0,
                "note": "unrecognised sheet, skipped",
            }

    total_inserted = sum(v["inserted"] for v in sheets_summary.values())
    total_skipped = sum(v["skipped"] for v in sheets_summary.values())

    return {
        "inserted": total_inserted,
        "skipped": total_skipped,
        "filename": filename,
        "uploaded_at": now.isoformat(),
        "sheets": sheets_summary,
    }
