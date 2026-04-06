import json
import hashlib
import io
from datetime import datetime, timezone
from fastapi import APIRouter, Depends, UploadFile, File
from typing import Annotated, Any
from fastapi.responses import StreamingResponse
from app.database import get_db
from app.dependencies import require_role
from app.core.errors import InvalidFileFormatError, ValidationError, NotFoundError
from app.models.contact import PreflightResult, ColumnMapping
from app.services.contact_parser import parse_contacts

router = APIRouter(prefix="/contacts", tags=["contacts"])

MAX_FILE_SIZE = 50 * 1024 * 1024  # 50 MB
RESULT_FILE_REF_KEY = "result.file_ref"


async def _cache_file_ref(file_ref: str, valid_rows: list) -> None:
    from redis.asyncio import from_url
    from app.config import settings
    from app.core.logging import get_logger

    logger = get_logger(__name__)

    try:
        redis = from_url(settings.redis_url, decode_responses=True)
        await redis.ping()  # Minimal connection check
        await redis.set(
            f"file_ref:{file_ref}",
            json.dumps([r.model_dump() for r in valid_rows]),
            ex=3600,
        )
        await redis.aclose()
    except Exception as e:
        logger.warning(
            "redis_cache_failed", 
            file_ref=file_ref, 
            error=str(e),
            detail="Contact file caching skipped. Fallback to DB will be used on campaign creation."
        )



@router.get("/template")
async def download_template(
    current_user: Annotated[dict, Depends(require_role("admin"))] = None,
):
    """Return a pre-formatted XLSX template the user can fill and re-upload."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contacts"

    # Header row
    headers = ["Name", "Email", "Phone"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="24422E")
        cell.alignment = Alignment(horizontal="center")

    # Sample rows so users know the expected format
    samples = [
        ("Jane Doe", "jane@example.com", "9820000001"),
        ("John Smith", "john@example.com", "9820000002"),
    ]
    for row_idx, (name, email, phone) in enumerate(samples, start=2):
        ws.cell(row=row_idx, column=1, value=name)
        ws.cell(row=row_idx, column=2, value=email)
        cell = ws.cell(row=row_idx, column=3, value=phone)
        cell.number_format = "@"

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=contacts_template.xlsx"},
    )


@router.post("/upload", response_model=PreflightResult)
async def upload_contacts(
    file: UploadFile = File(...),
    phone_column: str | None = None,
    email_column: str | None = None,
    name_column: str | None = None,
    current_user: Annotated[dict, Depends(require_role("admin"))] = None,
    db: Annotated[Any, Depends(get_db)] = None,
):
    if not file.filename.endswith((".xlsx", ".xls", ".csv")):
        raise InvalidFileFormatError("Only .xlsx, .xls, and .csv files are supported")

    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise ValidationError("File exceeds the 50 MB size limit")

    filename = file.filename
    file_hash = hashlib.sha256(content).hexdigest()
    uploader_id = str(current_user["_id"])

    # TEMPORARILY DISABLED CACHE TO DEBUG YOUR 0 RECIPIENTS ISSUE
    # existing = await db.contact_files.find_one(
    #     {"filename": filename, "hash": file_hash, "uploaded_by": uploader_id}
    # )
    # if existing:
    #     result = PreflightResult(**existing["result"])
    #     await _cache_file_ref(result.file_ref, result.valid_rows)
    #     return result

    mapping = ColumnMapping(
        phone_column=phone_column, 
        email_column=email_column, 
        name_column=name_column
    )

    suppressed = set()
    async for doc in db.suppression_list.find({}, {"phone": 1}):
        suppressed.add(doc["phone"])

    result = await parse_contacts(content, filename, mapping, suppressed)

    await db.contact_files.update_one(
        {"filename": filename, "hash": file_hash, "uploaded_by": uploader_id},
        {
            "$set": {
                "filename": filename,
                "hash": file_hash,
                "uploaded_by": uploader_id,
                "result": result.model_dump(),
                "uploaded_at": datetime.now(timezone.utc),
            }
        },
        upsert=True,
    )

    await _cache_file_ref(result.file_ref, result.valid_rows)
    return result


@router.get("/files")
async def list_contact_files(
    current_user: Annotated[dict, Depends(require_role("admin"))] = None,
    db: Annotated[Any, Depends(get_db)] = None,
):
    uploader_id = str(current_user["_id"])
    docs = (
        await db.contact_files.find(
            {"uploaded_by": uploader_id},
            {
                "filename": 1,
                "uploaded_at": 1,
                "result.valid_count": 1,
                "result.invalid_count": 1,
                RESULT_FILE_REF_KEY: 1,
            },
        )
        .sort("uploaded_at", -1)
        .to_list(100)
    )
    return [
        {
            "id": str(d["_id"]),
            "filename": d["filename"],
            "valid_count": d["result"]["valid_count"],
            "invalid_count": d["result"]["invalid_count"],
            "file_ref": d["result"]["file_ref"],
            "uploaded_at": d["uploaded_at"].isoformat(),
        }
        for d in docs
    ]


@router.delete("/files/{file_ref}", status_code=204)
async def delete_contact_file(
    file_ref: str,
    current_user: Annotated[dict, Depends(require_role("admin"))] = None,
    db: Annotated[Any, Depends(get_db)] = None,
):
    uploader_id = str(current_user["_id"])
    result = await db.contact_files.delete_one(
        {RESULT_FILE_REF_KEY: file_ref, "uploaded_by": uploader_id}
    )
    if result.deleted_count == 0:
        raise NotFoundError(f"Contact file '{file_ref}' not found")


@router.post("/files/{file_ref}/use", response_model=PreflightResult)
async def reuse_contact_file(
    file_ref: str,
    current_user: Annotated[dict, Depends(require_role("admin"))] = None,
    db: Annotated[Any, Depends(get_db)] = None,
):
    uploader_id = str(current_user["_id"])
    doc = await db.contact_files.find_one(
        {RESULT_FILE_REF_KEY: file_ref, "uploaded_by": uploader_id}
    )
    if not doc:
        raise NotFoundError(f"Contact file '{file_ref}' not found")

    result = PreflightResult(**doc["result"])
    await _cache_file_ref(result.file_ref, result.valid_rows)
    return result
