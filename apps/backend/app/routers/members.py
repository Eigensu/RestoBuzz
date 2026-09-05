"""Router for member management endpoints."""

import io
import json
import re
import uuid
import heapq

import openpyxl
from fastapi import APIRouter, Depends, Query, UploadFile, File, HTTPException
from redis.asyncio import from_url
from typing import Annotated, Any
from datetime import datetime, timezone, timedelta

from app.core.time import now_utc, normalize_external_dt
from app.services.member_match_service import member_match_service
from app.services import member_stats_service, member_segments

from app.config import settings
from app.database import get_db
from app.core.logging import get_logger
from app.core.utils import to_object_id
from pymongo.errors import DuplicateKeyError
from app.core.errors import (
    NotFoundError,
    ConflictError,
    ValidationError,
    InvalidFileFormatError,
)
from app.dependencies import (
    require_role,
    validate_restaurant_access,
    get_active_restaurant,
)
from app.models.member import (
    MemberCreate,
    MemberUpdate,
    MemberResponse,
    MemberListResponse,
)
from app.models.contact import PreflightResult, ContactRow, InvalidRow
from app.services.dormancy_service import dormancy_service, normalize_phone_for_match, DORMANCY_DAYS
from app.services.fielia_members_service import fielia_service, FieliaDatabaseError
from app.utils.phone import normalize_phone

router = APIRouter(prefix="/members", tags=["members"])
logger = get_logger(__name__)

# Constants for MongoDB operators to avoid duplication
REGEX = "$regex"
OPTIONS = "$options"

# How many rows the r2 hybrid path pulls from each source when a segment is
# active. Segments are derived per-row after mapping, so they cannot be pushed
# into the Fielia query — we over-fetch, then filter in memory.
R2_SEGMENT_FETCH_LIMIT = 10000


def _serialize(doc: dict, activity: tuple | None = None) -> MemberResponse:
    """Serialize a raw MongoDB member document to a MemberResponse."""
    last_visit = doc.get("last_visit")
    if last_visit and isinstance(last_visit, str):
        last_visit = datetime.fromisoformat(last_visit)

    last_visit_date, source = None, None
    if activity:
        last_visit_date, source = activity

    status, fallback_source = dormancy_service.compute_status(
        last_visit_date, last_visit
    )

    return MemberResponse(
        id=str(doc["_id"]),
        restaurant_id=doc.get("restaurant_id", "external"),
        type=doc.get("type", "nfc"),
        name=doc.get("name") or doc.get("guest_name") or "Unknown",
        phone=doc.get("phone") or doc.get("guest_number") or "Unknown",
        email=doc.get("email"),
        card_uid=doc.get("card_uid"),
        ecard_code=doc.get("ecard_code"),
        tags=doc.get("tags", []),
        notes=doc.get("notes"),
        visit_count=doc.get("visit_count", 0),
        last_visit=last_visit_date or last_visit,
        is_active=doc.get("is_active", True),
        normalized_phone=doc.get("normalized_phone"),
        dormancy_tier=doc.get("dormancy_tier", "UNKNOWN"),
        last_synced_at=doc.get("last_synced_at"),
        activity_status=status,
        activity_source=source or fallback_source,
        joined_at=doc.get("joined_at") or doc.get("created_at") or datetime.now(timezone.utc),
        interested_at=doc.get("interested_at"),
        interested_campaign_name=doc.get("interested_campaign_name"),
    )


async def _attach_message_stats(
    db: Any, restaurant_id: str, items: list[MemberResponse]
) -> list[MemberResponse]:
    """Merge each member's lifetime messaging rollup onto the response.

    One query for the whole page. Members with no rollup keep the zero defaults,
    which is the honest answer for someone we have never messaged.
    """
    if not items:
        return items
    stats_map = await member_stats_service.get_bulk_stats(
        db, restaurant_id, [i.phone for i in items]
    )
    for item in items:
        key = member_stats_service.phone_key(item.phone)
        member_stats_service.apply_stats(item, stats_map.get(key) if key else None)
    return items


async def _bulk_serialize(
    docs: list[dict], restaurant_id: str, db: Any, *, with_stats: bool = True
) -> list[MemberResponse]:
    """Serialize a batch of member docs with bulk activity lookup.

    `with_stats=False` skips the messaging rollup lookup — used by the r2 merge
    path, which over-fetches thousands of candidates and attaches stats once to
    the final page instead.
    """
    phones = [d.get("phone") for d in docs]
    uuids = [d.get("card_uid") for d in docs]
    activity_map = await dormancy_service.get_bulk_activity(
        db, restaurant_id, phones, uuids
    )
    items = []
    for doc in docs:
        norm_phone = normalize_phone_for_match(doc.get("phone"))
        uuid_val = doc.get("card_uid")
        activity = activity_map.get(uuid_val) or activity_map.get(norm_phone)
        items.append(_serialize(doc, activity))
    if not with_stats:
        return items
    return await _attach_message_stats(db, restaurant_id, items)


def _search_clause(search: str | None) -> dict:
    """Free-text clause over the fields a member is findable by.

    One definition for every listing path — the internal and r2 paths used to
    anchor `phone` differently, so the same query returned different results
    depending on which restaurant you were looking at.
    """
    if not search:
        return {}
    safe = re.escape(search)
    return {
        "$or": [
            {"name": {REGEX: safe, OPTIONS: "i"}},
            {"phone": {REGEX: safe, OPTIONS: "i"}},
            {"email": {REGEX: safe, OPTIONS: "i"}},
        ]
    }


@router.get("/segments")
async def list_segments() -> dict:
    """The behavioural segments this backend knows how to filter by.

    Served so the members page and the campaign audience picker render from the
    same list the query layer enforces, instead of each hardcoding their own.
    Categories are per-restaurant and come from the restaurant record.
    """
    return {"segments": member_segments.SEGMENT_DEFS}


@router.get("")
async def list_members(
    restaurant: Annotated[dict, Depends(get_active_restaurant)],
    member_type: Annotated[str | None, Query(alias="type")] = None,
    category: Annotated[str | None, Query()] = None,
    segment: Annotated[str | None, Query()] = None,
    search: Annotated[str | None, Query()] = None,
    page: Annotated[int, Query(ge=1)] = 1,
    page_size: Annotated[int, Query(ge=1, le=200)] = 50,
    db: Annotated[Any, Depends(get_db)] = None,
) -> MemberListResponse:
    rid = restaurant["id"]
    skip = (page - 1) * page_size
    category, segment = member_segments.resolve_axes(category, segment, member_type)

    if rid == "r2":
        return await _list_members_r2(db, category, segment, search, page, page_size)

    # 1. Resolve DB and Collection (handles all other restaurants)
    m_db, m_coll, m_filter = await member_match_service.get_member_db_context(rid)

    # 2. Build query from both axes. Category is matched without an allowlist,
    #    so an admin-defined category filters like any built-in one.
    query = member_segments.build_member_query(m_filter, category, segment)
    query.update(_search_clause(search))

    total = await m_db[m_coll].count_documents(query)
    docs = await (
        m_db[m_coll].find(query)
        .sort("joined_at", -1)
        .skip(skip)
        .limit(page_size)
        .to_list(length=page_size)
    )
    
    # 3. Serialize
    # We still use _bulk_serialize for legacy support/activity source info
    items = await _bulk_serialize(docs, rid, db)
    
    return MemberListResponse(items=items, total=total, page=page, page_size=page_size)


async def _list_members_r2(
    db: Any,
    category: str | None,
    segment: str | None,
    search: str | None,
    page: int,
    page_size: int,
) -> MemberListResponse:
    """Hybrid member listing for restaurant r2 (external Fielia + internal DB).

    Both sources are read, merged on joined_at, deduplicated by phone, then
    filtered and paginated. Filtering uses the same category/segment clauses as
    every other restaurant (member_segments), so "dormant" here means what it
    means everywhere else.

    Fielia stays read-only: we query it, map it, and never write back.
    """
    # A segment cannot be pushed into either source query: Fielia derives the
    # tier during mapping, so it has to be evaluated per row after the merge.
    # That is the only case that needs to over-fetch; category and search push
    # down cleanly, so an unsegmented view uses a rolling window.
    post_filtered = segment is not None
    fetch_limit = (
        R2_SEGMENT_FETCH_LIMIT if post_filtered else page * page_size + page_size * 3
    )

    # Fielia holds NFC cards only, so it can contribute to an unconstrained
    # listing or an explicit "nfc" request and nothing else.
    fielia_items: list[MemberResponse] = []
    fielia_total = 0
    if member_segments.fielia_supplies(category):
        fielia_res = await fielia_service.list_members(
            limit=fetch_limit,
            offset=0,
            search=search,
            member_type="nfc",
        )
        fielia_items = [MemberResponse(**item) for item in fielia_res["items"]]
        fielia_total = fielia_res["total"]

    internal_query = member_segments.build_member_query(
        {"restaurant_id": "r2"}, category, segment
    )
    internal_query.update(_search_clause(search))

    internal_total = await db.members.count_documents(internal_query)
    internal_docs = await (
        db.members.find(internal_query)
        .sort("joined_at", -1)
        .limit(fetch_limit)
        .to_list(length=fetch_limit)
    )
    internal_items = await _bulk_serialize(internal_docs, "r2", db, with_stats=False)

    # Merge two joined_at-descending streams in O(N).
    def sort_key(x):
        dt = x.joined_at or datetime.min
        # Normalize to UTC-aware so heapq can compare naive (Fielia) and
        # aware (internal MongoDB) datetimes without a TypeError.
        return (
            normalize_external_dt(dt)
            if dt != datetime.min
            else datetime.min.replace(tzinfo=timezone.utc)
        )

    merged_stream = heapq.merge(
        internal_items, fielia_items, key=sort_key, reverse=True
    )

    # Dedupe → filter → skip → collect.
    #
    # `total` was fielia_total + internal_total, which double-counted every
    # member present in both sources AND ignored the segment filter entirely —
    # so on a filtered view the page count and the Next button were wrong by
    # however many members the filter excluded.
    #
    # Now: when we post-filter we are already walking the whole (capped) merged
    # stream, so we count exactly what survives. When we don't, the filters ran
    # inside both source queries, so their own counts are authoritative and we
    # can stop as soon as the page is full — that sum still counts a member
    # present in both sources twice, which is why it is labelled an estimate.
    seen_phones: set[str] = set()
    results: list[MemberResponse] = []
    skip_count = (page - 1) * page_size
    matched = 0

    for item in merged_stream:
        phone = normalize_phone_for_match(item.phone) or f"no_phone_{item.id}"
        if phone in seen_phones:
            continue
        seen_phones.add(phone)

        if not member_segments.matches_category(item, category):
            continue
        if not member_segments.matches_segment(item, segment):
            continue

        matched += 1
        if skip_count > 0:
            skip_count -= 1
            continue
        if len(results) < page_size:
            results.append(item)
        elif not post_filtered:
            # Page is full and the count comes from the sources, so there is
            # nothing left to learn from the rest of the stream.
            break

    # Messaging rollups are attached only to the page we actually return —
    # both source streams over-fetch by design, and stats for rows that get
    # filtered out would be thrown away.
    await _attach_message_stats(db, "r2", results)

    total = matched if post_filtered else fielia_total + internal_total

    return MemberListResponse(
        items=results,
        total=total,
        page=page,
        page_size=page_size,
    )


@router.post("", status_code=201)
async def create_member(
    body: MemberCreate,
    restaurant: Annotated[dict, Depends(get_active_restaurant)],
    _user: Annotated[dict, Depends(require_role("admin"))],
    db: Annotated[Any, Depends(get_db)],
) -> MemberResponse:
    """Create a new member for the active restaurant."""
    valid_categories = restaurant.get("member_categories") or ["nfc", "ecard"]
    if body.type not in valid_categories:
        raise ValidationError(
            f"Invalid member type '{body.type}'. "
            f"Valid types: {', '.join(valid_categories)}"
        )

    if body.type == "nfc" and not body.card_uid:
        raise ValidationError("card_uid is required for NFC members")
    if body.type == "ecard" and not body.ecard_code:
        raise ValidationError("ecard_code is required for e-card members")

    # Only check uniqueness when a real phone is provided
    if body.phone:
        existing = await db.members.find_one(
            {"restaurant_id": restaurant["id"], "phone": body.phone}
        )
        if existing:
            raise ConflictError(
                "A member with this phone number already exists in our internal database"
            )

        if restaurant["id"] == "r2":
            try:
                if await fielia_service.check_phone_exists(body.phone):
                    raise ConflictError(
                        "A member with this phone number already exists in the Fielia database"
                    )
            except FieliaDatabaseError as exc:
                raise HTTPException(
                    status_code=503,
                    detail="External member service unavailable. Please try again later.",
                    headers={"Retry-After": "10"}
                ) from exc

    now = now_utc()
    doc = {
        "restaurant_id": restaurant["id"],
        "type": body.type,
        "name": body.name,
        "phone": body.phone or None,
        "email": body.email,
        "card_uid": body.card_uid,
        "ecard_code": body.ecard_code,
        "tags": body.tags,
        "notes": body.notes,
        "visit_count": 0,
        "last_visit": None,
        "is_active": True,
        "joined_at": now,
    }
    try:
        result = await db.members.insert_one(doc)
    except DuplicateKeyError:
        raise ConflictError("A member with this phone number already exists (concurrent write detected)")

    doc["_id"] = result.inserted_id
    return _serialize(doc)


@router.get("/{member_id}")
async def get_member(
    member_id: str,
    current_user: Annotated[dict, Depends(require_role("viewer"))],
    db: Annotated[Any, Depends(get_db)],
) -> MemberResponse:
    """Fetch a single member by ID."""
    doc = await db.members.find_one({"_id": to_object_id(member_id)})
    if not doc:
        raise NotFoundError(f"Member '{member_id}' not found")
    await validate_restaurant_access(current_user, doc["restaurant_id"], db)
    item = _serialize(doc)
    await _attach_message_stats(db, doc["restaurant_id"], [item])
    return item


@router.patch("/{member_id}")
async def update_member(
    member_id: str,
    body: MemberUpdate,
    current_user: Annotated[dict, Depends(require_role("admin"))],
    db: Annotated[Any, Depends(get_db)],
) -> MemberResponse:
    """Update fields on an existing member."""
    doc = await db.members.find_one({"_id": to_object_id(member_id)})
    if not doc:
        raise NotFoundError(f"Member '{member_id}' not found")
    await validate_restaurant_access(current_user, doc["restaurant_id"], db)

    updates = {k: v for k, v in body.model_dump().items() if v is not None}
    if not updates:
        raise ValidationError("No fields provided to update")

    doc = await db.members.find_one_and_update(
        {"_id": to_object_id(member_id)},
        {"$set": updates},
        return_document=True,
    )
    item = _serialize(doc)
    await _attach_message_stats(db, doc["restaurant_id"], [item])
    return item


@router.delete("/{member_id}", status_code=204)
async def delete_member(
    member_id: str,
    current_user: Annotated[dict, Depends(require_role("admin"))],
    db: Annotated[Any, Depends(get_db)],
) -> None:
    """Delete a member by ID."""
    doc = await db.members.find_one({"_id": to_object_id(member_id)})
    if not doc:
        raise NotFoundError(f"Member '{member_id}' not found")
    await validate_restaurant_access(current_user, doc["restaurant_id"], db)
    await db.members.delete_one({"_id": to_object_id(member_id)})


@router.post("/{member_id}/visit")
async def record_visit(
    member_id: str,
    current_user: Annotated[dict, Depends(require_role("admin"))],
    db: Annotated[Any, Depends(get_db)],
) -> MemberResponse:
    """Increment visit count and update last_visit timestamp for a member."""
    doc = await db.members.find_one({"_id": to_object_id(member_id)})
    if not doc:
        raise NotFoundError(f"Member '{member_id}' not found")
    await validate_restaurant_access(current_user, doc["restaurant_id"], db)

    now = now_utc()
    doc = await db.members.find_one_and_update(
        {"_id": to_object_id(member_id)},
        {"$inc": {"visit_count": 1}, "$set": {"last_visit": now}},
        return_document=True,
    )
    item = _serialize(doc)
    await _attach_message_stats(db, doc["restaurant_id"], [item])
    return item


@router.post("/{member_id}/send-ecard")
async def send_member_ecard(
    member_id: str,
    current_user: Annotated[dict, Depends(require_role("admin"))],
    db: Annotated[Any, Depends(get_db)],
) -> dict:
    """Send a personalized membership e-card to a single member.

    Renders the member's name onto the restaurant's configured base card and
    sends it via the restaurant's WABA using the configured e-card template.
    Reuses the same send path as campaigns, so delivery/read webhooks apply.
    """
    doc = await db.members.find_one({"_id": to_object_id(member_id)})
    if not doc:
        raise NotFoundError(f"Member '{member_id}' not found")
    restaurant_id = doc["restaurant_id"]
    await validate_restaurant_access(current_user, restaurant_id, db)

    name = (doc.get("name") or "").strip()
    phone = normalize_phone(doc.get("phone") or "")
    if not phone:
        raise ValidationError("This member has no valid phone number.")
    if not name:
        raise ValidationError(
            "This member has no name — the e-card needs a name to render."
        )

    rest = await db.restaurants.find_one({"id": restaurant_id})
    if not rest:
        # Some member docs store the restaurant's _id string instead of its
        # slug-style "id"; fall back to the ObjectId lookup like elsewhere.
        from bson import ObjectId
        from bson.errors import InvalidId

        try:
            rest = await db.restaurants.find_one({"_id": ObjectId(restaurant_id)})
        except (InvalidId, TypeError):
            rest = None
    cfg = (rest or {}).get("ecard_config")
    if not cfg or not cfg.get("base_public_id") or not cfg.get("template_name"):
        raise ValidationError(
            "This restaurant has no e-card configured. Set up the e-card base "
            "image and template first."
        )

    from app.services.ecard_service import build_card_url
    from app.services.meta_api import send_template_message, MetaAPIError
    from app.services.campaign_service import resolve_waba_credentials

    media_url = build_card_url(cfg["base_public_id"], name, cfg.get("overlay") or {})
    wa_phone_id, wa_access_token, _ = await resolve_waba_credentials(db, restaurant_id)

    try:
        wa_message_id, endpoint_used = await send_template_message(
            to=phone,
            template_name=cfg["template_name"],
            variables={},
            media_url=media_url,
            language=cfg.get("language") or "en",
            phone_id=wa_phone_id,
            access_token=wa_access_token,
            media_type="image",
        )
    except MetaAPIError as e:
        raise ValidationError(f"WhatsApp rejected the e-card: {e.message}") from e

    await db.outbound_messages.insert_one(
        {
            "wa_message_id": wa_message_id,
            "to_phone": phone,
            "body": f"E-card sent to {name}",
            "status": "sent",
            "sent_at": now_utc(),
            "restaurant_id": restaurant_id,
            "wa_phone_id": wa_phone_id,
            "sender_name": "System (E-card)",
            "channel": "whatsapp",
            "media_url": media_url,
        }
    )
    logger.info(
        "member_ecard_sent",
        member_id=member_id,
        restaurant_id=restaurant_id,
        wa_message_id=wa_message_id,
    )
    return {"wa_message_id": wa_message_id, "endpoint_used": endpoint_used}


@router.post("/as-contacts")
async def members_as_contacts(
    restaurant: Annotated[dict, Depends(get_active_restaurant)],
    _user: Annotated[dict, Depends(require_role("admin"))],
    db: Annotated[Any, Depends(get_db)] = None,
    member_type: Annotated[str | None, Query(alias="type")] = None,
    category: Annotated[str | None, Query()] = None,
    segment: Annotated[str | None, Query()] = None,
    limit: Annotated[int | None, Query(ge=1)] = None,
) -> PreflightResult:
    """
    Convert members into a PreflightResult for use as campaign contacts.

    Audience is picked on the same two axes as the members listing — an
    optional category and an optional segment — so a campaign targets exactly
    the people the corresponding tab shows.

    Sources:
    - reservego: combines reservego_uploads + reservego_bill_data collections
    - r2 (Fielia): external Fielia NFC members plus internal DB members
    - all others: queries internal members DB only
    """
    if member_type != "reservego":
        category, segment = member_segments.resolve_axes(
            category, segment, member_type
        )

    suppressed: set[str] = set()
    async for sup in db.suppression_list.find({}, {"phone": 1}):
        suppressed.add(sup["phone"])

    valid_rows: list[ContactRow] = []
    invalid_rows: list[InvalidRow] = []
    seen_phones: set[str] = set()
    duplicate_count = 0
    suppressed_count = 0
    row_num = 1

    def process_row(name: str, raw_phone_val: Any) -> None:
        nonlocal row_num, duplicate_count, suppressed_count
        raw_phone = str(raw_phone_val).strip() if raw_phone_val else ""
        if not raw_phone:
            invalid_rows.append(
                InvalidRow(row_number=row_num, raw_value="", reason="Empty phone")
            )
            row_num += 1
            return
        normalized = normalize_phone(raw_phone)
        if not normalized:
            invalid_rows.append(
                InvalidRow(
                    row_number=row_num,
                    raw_value=raw_phone,
                    reason="Invalid phone number",
                )
            )
            row_num += 1
            return
        if normalized in seen_phones:
            duplicate_count += 1
            row_num += 1
            return
        seen_phones.add(normalized)
        if normalized in suppressed:
            suppressed_count += 1
            row_num += 1
            return
        valid_rows.append(ContactRow(name=name or "", phone=normalized, variables={}))
        row_num += 1

    if member_type == "reservego" or category == "reservego":
        await _process_reservego(db, restaurant["id"], limit, process_row, valid_rows)
    else:
        await _process_members(
            db, restaurant, category, segment, limit, process_row, valid_rows
        )

    file_ref = str(uuid.uuid4())
    redis = from_url(settings.redis_url, decode_responses=True)
    await redis.set(
        f"file_ref:{file_ref}",
        json.dumps([r.model_dump() for r in valid_rows]),
        ex=3600,
    )
    await redis.aclose()

    return PreflightResult(
        valid_count=len(valid_rows),
        invalid_count=len(invalid_rows),
        duplicate_count=duplicate_count,
        suppressed_count=suppressed_count,
        valid_rows=valid_rows,
        invalid_rows=invalid_rows,
        file_ref=file_ref,
    )


async def _process_reservego(
    db: Any,
    restaurant_id: str,
    limit: int | None,
    process_row: Any,
    valid_rows: list,
) -> None:
    """Stream ReserveGo uploads and bill data into the contact processor."""
    async for doc in db.reservego_uploads.find(
        {"restaurant_id": restaurant_id}, {"guest_name": 1, "phone": 1}
    ).sort("_id", -1):
        if limit and len(valid_rows) >= limit:
            return
        process_row(doc.get("guest_name", ""), doc.get("phone"))

    async for doc in db.reservego_bill_data.find(
        {"restaurant_id": restaurant_id}, {"guest_name": 1, "guest_number": 1}
    ).sort("_id", -1):
        if limit and len(valid_rows) >= limit:
            return
        process_row(doc.get("guest_name", ""), doc.get("guest_number"))


async def _process_fielia_members(
    category: str | None,
    segment: str | None,
    limit: int | None,
    process_row: Any,
    valid_rows: list,
) -> None:
    """Stream external Fielia members into the contact processor.

    Read-only against Fielia. Segments are evaluated on the mapped row because
    the tier is derived during mapping, not stored in the external dataset.
    """
    if not member_segments.fielia_supplies(category):
        return
    fielia_res = await fielia_service.list_members(
        limit=R2_SEGMENT_FETCH_LIMIT, offset=0, member_type="nfc"
    )
    for m in fielia_res["items"]:
        if limit and len(valid_rows) >= limit:
            return
        if segment == "inactive":
            if m.get("dormancy_tier") not in member_segments.INACTIVE_TIERS:
                continue
        elif segment in ("active", "at_risk", "dormant", "lost"):
            if m.get("dormancy_tier") != segment.upper():
                continue
        elif segment == "interested":
            # Fielia carries no campaign tags; interested members live in our DB.
            continue
        process_row(m.get("name", "Unknown"), m.get("phone"))


async def _process_members(
    db: Any,
    restaurant: dict,
    category: str | None,
    segment: str | None,
    limit: int | None,
    process_row: Any,
    valid_rows: list,
) -> None:
    """Stream members into the contact processor for a category/segment.

    Uses the same filter builder as the members listing, so a campaign audience
    and the tab it was picked from resolve to the same people.
    """
    rid = restaurant["id"]

    # r2 draws from Fielia *and* the internal DB. The internal half used to be
    # skipped entirely, so members added in-app were never campaign targets.
    if rid == "r2":
        await _process_fielia_members(
            category, segment, limit, process_row, valid_rows
        )
        internal_query = member_segments.build_member_query(
            {"restaurant_id": "r2", "is_active": True}, category, segment
        )
        async for doc in (
            db.members.find(internal_query, {"name": 1, "phone": 1}).sort("_id", -1)
        ):
            if limit and len(valid_rows) >= limit:
                return
            process_row(doc.get("name", ""), doc.get("phone"))
        return

    m_db, m_coll, m_filter = await member_match_service.get_member_db_context(rid)
    query = member_segments.build_member_query(
        {**m_filter, "is_active": True}, category, segment
    )

    async for doc in m_db[m_coll].find(query, {"name": 1, "phone": 1}).sort("_id", -1):
        if limit and len(valid_rows) >= limit:
            return
        process_row(doc.get("name", ""), doc.get("phone"))


@router.post("/import")
async def import_members(
    restaurant: Annotated[dict, Depends(get_active_restaurant)],
    file: Annotated[UploadFile, File()],
    _user: Annotated[dict, Depends(require_role("admin"))],
    db: Annotated[Any, Depends(get_db)],
    member_type: Annotated[str, Query(alias="type")] = "ecard",
) -> dict:
    """Bulk-import members from an uploaded XLSX file."""
    # The import type is a category, never a segment. Without this check the
    # members page could import from a segment tab and stamp every row with a
    # type ("inactive", "interested") that no category tab can ever show.
    valid_categories = restaurant.get("member_categories") or ["nfc", "ecard"]
    if member_type not in valid_categories:
        raise ValidationError(
            f"Invalid member type '{member_type}'. "
            f"Valid types: {', '.join(valid_categories)}"
        )

    filename = file.filename or ""
    content_type = file.content_type or ""
    allowed_content_types = {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/octet-stream",
        "application/vnd.ms-excel",
    }
    if not (
        content_type in allowed_content_types or filename.lower().endswith(".xlsx")
    ):
        logger.error(
            "import_invalid_format", content_type=content_type, filename=filename
        )
        raise InvalidFileFormatError("Only .xlsx Excel files are supported for import")

    contents = await file.read()
    if not contents:
        raise InvalidFileFormatError("Uploaded Excel file is empty")
    if len(contents) > 10 * 1024 * 1024:
        raise ValidationError("Excel file is too large (max 10MB)")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True)
    except (ValueError, KeyError) as exc:
        raise InvalidFileFormatError("Unable to read Excel file") from exc

    ws = wb.active
    if ws.max_row and ws.max_row > 5001:
        raise ValidationError("Excel file has too many rows (max 5000)")

    raw_headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]

    def find_col(names: list[str]) -> int | None:
        for n in names:
            if n in raw_headers:
                return raw_headers.index(n)
        return None

    name_idx = find_col(["name", "full name", "fullname", "customer name"])
    phone_idx = find_col(
        ["phone", "contact number", "mobile", "phone number", "contact"]
    )
    email_idx = find_col(["email", "email address"])
    card_uid_idx = find_col(
        ["card_uid", "card id", "uid", "card number", "card nfc id"]
    )
    ecard_code_idx = find_col(["ecard_code", "ecard code", "e-card code", "code"])

    if name_idx is None:
        raise InvalidFileFormatError("Excel must have a 'Name' column")

    now = now_utc()
    inserted = skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        name = str(row[name_idx]).strip() if row[name_idx] else ""
        if not name:
            skipped += 1
            continue

        raw_phone = (
            str(row[phone_idx]).strip()
            if phone_idx is not None and row[phone_idx]
            else ""
        )
        email = (
            str(row[email_idx]).strip()
            if email_idx is not None and row[email_idx]
            else None
        )
        card_uid = (
            str(row[card_uid_idx]).strip()
            if card_uid_idx is not None and row[card_uid_idx]
            else None
        )
        ecard_code = (
            str(row[ecard_code_idx]).strip()
            if ecard_code_idx is not None and row[ecard_code_idx]
            else None
        )

        if raw_phone and raw_phone != "None":
            phone = normalize_phone(raw_phone)
            if phone is None:
                skipped += 1
                continue
        else:
            phone = None  # store None, not "", to avoid duplicate-empty-phone index collisions

        if phone:
            existing = await db.members.find_one(
                {"restaurant_id": restaurant["id"], "phone": phone}
            )
            if existing:
                skipped += 1
                continue

        await db.members.insert_one(
            {
                "restaurant_id": restaurant["id"],
                "type": member_type,
                "name": name,
                "phone": phone,
                "email": email,
                "card_uid": card_uid,
                "ecard_code": ecard_code,
                "tags": [],
                "notes": None,
                "visit_count": 0,
                "last_visit": None,
                "is_active": True,
                "joined_at": now,
                "source": "excel",
            }
        )
        inserted += 1

    return {"inserted": inserted, "skipped": skipped}


@router.delete("/bulk", status_code=204)
async def bulk_delete_members(
    restaurant: Annotated[dict, Depends(get_active_restaurant)],
    _user: Annotated[dict, Depends(require_role("admin"))],
    source: Annotated[str | None, Query()] = None,
    delete_all: Annotated[bool, Query(alias="deleteAll")] = False,
    db: Annotated[Any, Depends(get_db)] = None,
) -> None:
    """Bulk delete members. Requires either deleteAll=true or a source filter."""
    logger.info(
        "bulk_delete_request",
        restaurant_id=restaurant["id"],
        source=source,
        delete_all=delete_all,
    )

    query: dict = {"restaurant_id": restaurant["id"]}
    if delete_all:
        pass  # no additional filter
    elif source:
        query["source"] = source
    else:
        raise ValidationError(
            "Must specify either 'deleteAll=true' or a 'source' to delete in bulk"
        )

    result = await db.members.delete_many(query)
    logger.info("bulk_delete_result", deleted_count=result.deleted_count)
